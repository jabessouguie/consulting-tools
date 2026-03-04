"""
Tests pour TenderScout AI :
  - TenderScoutAgent (scraping + analyse Gemini)
  - TenderDatabase (SQLite idempotente)
  - Endpoints FastAPI (/tenderscout, /api/tenderscout/*)
"""

import json
import os
import tempfile
from io import BytesIO
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest
from fastapi.testclient import TestClient

# ── Client HTTP de test ───────────────────────────────────────────────
with (
    patch("utils.llm_client.LLMClient.__init__", return_value=None),
    patch("agents.tender_scout_agent.LLMClient"),
    patch("app.get_consultant_info", return_value={"name": "Test Consultant"}),
):
    from app import app

client = TestClient(app)

# ── Helpers ───────────────────────────────────────────────────────────
CSRF_HEADERS = {"origin": "http://testserver"}

MOCK_USER = {"username": "test", "name": "Test User"}


def _mock_user(request):
    return MOCK_USER


BOAMP_JSON_RESPONSE = {
    "results": [
        {
            "idweb": "2024AO001",
            "objet": "Mission de conseil en transformation digitale",
            "acheteur": "Ministère de l'Économie",
            "dateparution": "2024-03-01",
            "datelimitereponse": "2024-04-15",
            "descripteur": "Conseil en transformation digitale et conduite du changement",
        },
        {
            "idweb": "2024AO002",
            "objet": "Audit organisationnel",
            "dateparution": "2024-03-02",
            "datelimitereponse": "2024-04-20",
        },
    ]
}

FRANCEMARCHES_HTML = """
<html><body>
<article class="avis-list-item">
  <a href="/avis/fm-12345">
    <h3 class="title">Consulting IT et transformation digitale</h3>
  </a>
  <span class="buyer">Conseil Régional Île-de-France</span>
  <time class="date" datetime="2024-03-05">5 mars 2024</time>
  <div class="excerpt">Prestation de conseil en stratégie digitale</div>
</article>
</body></html>
"""

GEMINI_ANALYSIS = {
    "decision": "GO",
    "score": 82,
    "resume": "Mission de conseil stratégique adaptée à nos compétences.",
    "budget_estime": "200 000 € HT",
    "echeance": "15 avril 2024",
    "criteres_notation": ["Expérience sectorielle", "Méthodologie"],
    "risques": ["Délai court"],
    "atouts": ["Forte expérience en transformation digitale"],
    "recommandation": "Répondre prioritairement.",
}

TENDER_DATA = {
    "reference": "BOAMP-2024AO001",
    "titre": "Mission de conseil en transformation digitale",
    "acheteur": "Ministère de l'Économie",
    "source": "boamp",
    "url": "https://www.boamp.fr/avis/detail/2024AO001",
    "date_publication": "2024-03-01",
    "date_limite": "2024-04-15",
    "description": "Conseil en transformation digitale",
}


# ============================================================
# TestTenderScoutAgent
# ============================================================


class TestTenderScoutAgent:
    """Tests pour TenderScoutAgent (scraping + analyse)."""

    def setup_method(self):
        with patch("agents.tender_scout_agent.LLMClient"):
            from agents.tender_scout_agent import TenderScoutAgent

            self.agent = TenderScoutAgent.__new__(TenderScoutAgent)
            self.agent.llm = MagicMock()

    # ── scrape_boamp ─────────────────────────────────────────

    def test_scrape_boamp_ok(self):
        mock_resp = MagicMock()
        mock_resp.raise_for_status.return_value = None
        mock_resp.json.return_value = BOAMP_JSON_RESPONSE

        with patch("agents.tender_scout_agent.requests.get", return_value=mock_resp):
            tenders = self.agent.scrape_boamp(["transformation digitale"])

        assert len(tenders) == 2
        assert tenders[0]["reference"] == "BOAMP-2024AO001"
        assert tenders[0]["source"] == "boamp"
        assert tenders[0]["titre"] == "Mission de conseil en transformation digitale"
        assert "boamp.fr" in tenders[0]["url"]

    def test_scrape_boamp_multiple_keywords(self):
        """Deux mots-clés → deux appels HTTP, résultats distincts mergés."""
        mock_resp1 = MagicMock()
        mock_resp1.raise_for_status.return_value = None
        mock_resp1.json.return_value = {
            "results": [
                {"idweb": "AAA001", "objet": "Mission A", "dateparution": "2024-01-01"}
            ]
        }
        mock_resp2 = MagicMock()
        mock_resp2.raise_for_status.return_value = None
        mock_resp2.json.return_value = {
            "results": [
                {"idweb": "BBB001", "objet": "Mission B", "dateparution": "2024-01-02"}
            ]
        }

        with patch(
            "agents.tender_scout_agent.requests.get",
            side_effect=[mock_resp1, mock_resp2],
        ):
            tenders = self.agent.scrape_boamp(["data", "conseil"])

        assert len(tenders) == 2
        refs = {t["reference"] for t in tenders}
        assert "BOAMP-AAA001" in refs
        assert "BOAMP-BBB001" in refs

    def test_scrape_boamp_deduplicates(self):
        """Même référence dans deux keywords → une seule entrée."""
        mock_resp = MagicMock()
        mock_resp.raise_for_status.return_value = None
        mock_resp.json.return_value = {
            "results": [{"idweb": "DUP001", "objet": "Doublon"}]
        }

        with patch(
            "agents.tender_scout_agent.requests.get",
            side_effect=[mock_resp, mock_resp],
        ):
            tenders = self.agent.scrape_boamp(["word1", "word2"])

        assert len(tenders) == 1

    def test_scrape_boamp_http_error(self):
        import requests as req

        from agents.tender_scout_agent import ScrapingError

        mock_resp = MagicMock()
        mock_resp.raise_for_status.side_effect = req.HTTPError("404")

        with patch("agents.tender_scout_agent.requests.get", return_value=mock_resp):
            with pytest.raises(ScrapingError, match="BOAMP"):
                self.agent.scrape_boamp(["test"])

    def test_scrape_boamp_invalid_json(self):
        from agents.tender_scout_agent import ScrapingError

        mock_resp = MagicMock()
        mock_resp.raise_for_status.return_value = None
        mock_resp.json.side_effect = ValueError("not json")

        with patch("agents.tender_scout_agent.requests.get", return_value=mock_resp):
            with pytest.raises(ScrapingError, match="non-JSON"):
                self.agent.scrape_boamp(["test"])

    def test_scrape_boamp_connection_error(self):
        import requests as req

        from agents.tender_scout_agent import ScrapingError

        with patch(
            "agents.tender_scout_agent.requests.get",
            side_effect=req.ConnectionError("timeout"),
        ):
            with pytest.raises(ScrapingError):
                self.agent.scrape_boamp(["test"])

    # ── scrape_francemarches ─────────────────────────────────

    def test_scrape_francemarches_ok(self):
        mock_resp = MagicMock()
        mock_resp.raise_for_status.return_value = None
        mock_resp.text = FRANCEMARCHES_HTML

        with patch("agents.tender_scout_agent.requests.get", return_value=mock_resp):
            tenders = self.agent.scrape_francemarches(["transformation digitale"])

        assert len(tenders) >= 1
        assert tenders[0]["source"] == "francemarches"
        assert "Consulting IT" in tenders[0]["titre"]

    def test_scrape_francemarches_http_error(self):
        import requests as req

        from agents.tender_scout_agent import ScrapingError

        mock_resp = MagicMock()
        mock_resp.raise_for_status.side_effect = req.HTTPError("500")

        with patch("agents.tender_scout_agent.requests.get", return_value=mock_resp):
            with pytest.raises(ScrapingError, match="Francemarches"):
                self.agent.scrape_francemarches(["test"])

    def test_scrape_francemarches_connection_error(self):
        import requests as req

        from agents.tender_scout_agent import ScrapingError

        with patch(
            "agents.tender_scout_agent.requests.get",
            side_effect=req.ConnectionError("refused"),
        ):
            with pytest.raises(ScrapingError):
                self.agent.scrape_francemarches(["test"])

    def test_scrape_francemarches_empty_page(self):
        """Page sans résultats → liste vide."""
        mock_resp = MagicMock()
        mock_resp.raise_for_status.return_value = None
        mock_resp.text = "<html><body><p>Aucun résultat</p></body></html>"

        with patch("agents.tender_scout_agent.requests.get", return_value=mock_resp):
            tenders = self.agent.scrape_francemarches(["xyz_no_match"])

        assert tenders == []

    # ── analyze_tender ───────────────────────────────────────

    def test_analyze_tender_go(self):
        self.agent.llm.extract_structured_data.return_value = dict(GEMINI_ANALYSIS)

        result = self.agent.analyze_tender(TENDER_DATA)

        assert result["decision"] == "GO"
        assert result["score"] == 82
        assert "resume" in result

    def test_analyze_tender_nogo(self):
        analysis = dict(GEMINI_ANALYSIS)
        analysis["decision"] = "NO-GO"
        analysis["score"] = 15
        self.agent.llm.extract_structured_data.return_value = analysis

        result = self.agent.analyze_tender(TENDER_DATA)

        assert result["decision"] == "NO-GO"
        assert result["score"] == 15

    def test_analyze_tender_a_etudier(self):
        analysis = dict(GEMINI_ANALYSIS)
        analysis["decision"] = "A_ETUDIER"
        analysis["score"] = 50
        self.agent.llm.extract_structured_data.return_value = analysis

        result = self.agent.analyze_tender(TENDER_DATA)

        assert result["decision"] == "A_ETUDIER"

    def test_analyze_tender_normalizes_unknown_decision(self):
        """Une décision inconnue est normalisée en A_ETUDIER."""
        analysis = dict(GEMINI_ANALYSIS)
        analysis["decision"] = "MAYBE"
        self.agent.llm.extract_structured_data.return_value = analysis

        result = self.agent.analyze_tender(TENDER_DATA)

        assert result["decision"] == "A_ETUDIER"

    def test_analyze_tender_normalizes_score(self):
        """Score non-entier → converti en int."""
        analysis = dict(GEMINI_ANALYSIS)
        analysis["score"] = "75"
        self.agent.llm.extract_structured_data.return_value = analysis

        result = self.agent.analyze_tender(TENDER_DATA)

        assert result["score"] == 75

    def test_analyze_tender_invalid_score_defaults_to_zero(self):
        analysis = dict(GEMINI_ANALYSIS)
        analysis["score"] = "invalid"
        self.agent.llm.extract_structured_data.return_value = analysis

        result = self.agent.analyze_tender(TENDER_DATA)

        assert result["score"] == 0

    def test_analyze_tender_llm_exception(self):
        from agents.tender_scout_agent import AnalysisError

        self.agent.llm.extract_structured_data.side_effect = RuntimeError("API down")

        with pytest.raises(AnalysisError, match="LLM"):
            self.agent.analyze_tender(TENDER_DATA)

    def test_analyze_tender_empty_result(self):
        from agents.tender_scout_agent import AnalysisError

        self.agent.llm.extract_structured_data.return_value = {}

        with pytest.raises(AnalysisError, match="vide ou invalide"):
            self.agent.analyze_tender(TENDER_DATA)

    def test_analyze_tender_none_result(self):
        from agents.tender_scout_agent import AnalysisError

        self.agent.llm.extract_structured_data.return_value = None

        with pytest.raises(AnalysisError):
            self.agent.analyze_tender(TENDER_DATA)

    def test_init_creates_llm_client(self):
        """Le constructeur instancie LLMClient (couvre la ligne d'init)."""
        with patch("agents.tender_scout_agent.LLMClient") as MockLLM:
            from agents.tender_scout_agent import TenderScoutAgent

            agent = TenderScoutAgent(api_key="mykey", model="gemini-flash")

        MockLLM.assert_called_once_with(
            api_key="mykey", model="gemini-flash"
        )
        assert agent.llm is MockLLM.return_value

    def test_scrape_francemarches_limit_respected(self):
        """Avec limit=1, seul le premier AO est retourné (couvre le `break`)."""
        html = """<html><body>
        <article class="avis-list-item">
          <a href="/avis/one"><h3>Mission 1</h3></a>
        </article>
        <article class="avis-list-item">
          <a href="/avis/two"><h3>Mission 2</h3></a>
        </article>
        </body></html>"""
        mock_resp = MagicMock()
        mock_resp.raise_for_status.return_value = None
        mock_resp.text = html

        with patch("agents.tender_scout_agent.requests.get", return_value=mock_resp):
            tenders = self.agent.scrape_francemarches(["test"], limit=1)

        assert len(tenders) == 1

    def test_scrape_francemarches_skips_card_without_link(self):
        """Card sans <a> → ignorée (couvre la ligne `continue` après link_tag check)."""
        html = """<html><body>
        <article class="avis-list-item">
          <h3>Mission sans lien</h3>
        </article>
        <article class="avis-list-item">
          <a href="/avis/with-link"><h3>Mission avec lien</h3></a>
        </article>
        </body></html>"""
        mock_resp = MagicMock()
        mock_resp.raise_for_status.return_value = None
        mock_resp.text = html

        with patch("agents.tender_scout_agent.requests.get", return_value=mock_resp):
            tenders = self.agent.scrape_francemarches(["test"])

        # Seule la card avec lien est retournée
        assert len(tenders) == 1
        assert "with-link" in tenders[0]["url"]

    def test_scrape_francemarches_deduplicates_urls(self):
        """Deux cards avec la même URL → une seule entrée (couvre le `continue` sur seen_urls)."""
        html = """<html><body>
        <article class="avis-list-item">
          <a href="/avis/same-url"><h3>Mission A</h3></a>
        </article>
        <article class="avis-list-item">
          <a href="/avis/same-url"><h3>Mission B (doublon)</h3></a>
        </article>
        </body></html>"""
        mock_resp = MagicMock()
        mock_resp.raise_for_status.return_value = None
        mock_resp.text = html

        with patch("agents.tender_scout_agent.requests.get", return_value=mock_resp):
            tenders = self.agent.scrape_francemarches(["test"])

        assert len(tenders) == 1

    def test_scrape_francemarches_fallback_sans_titre(self):
        """Lien sans texte → titre = 'Sans titre' (couvre la ligne de fallback)."""
        html = """<html><body>
        <article class="avis-list-item">
          <a href="/avis/no-text"></a>
        </article>
        </body></html>"""
        mock_resp = MagicMock()
        mock_resp.raise_for_status.return_value = None
        mock_resp.text = html

        with patch("agents.tender_scout_agent.requests.get", return_value=mock_resp):
            tenders = self.agent.scrape_francemarches(["test"])

        assert len(tenders) == 1
        assert tenders[0]["titre"] == "Sans titre"


# ============================================================
# TestTenderDatabase
# ============================================================


class TestTenderDatabase:
    """Tests pour TenderDatabase (SQLite)."""

    @pytest.fixture(autouse=True)
    def db(self, tmp_path):
        from utils.tender_db import TenderDatabase

        db_path = str(tmp_path / "test_tenders.db")
        self._db = TenderDatabase(db_path=db_path)
        return self._db

    def test_init_creates_table(self, tmp_path):
        import sqlite3

        from utils.tender_db import TenderDatabase

        db_path = str(tmp_path / "init_test.db")
        TenderDatabase(db_path=db_path)

        with sqlite3.connect(db_path) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
            tables = {row[0] for row in cursor.fetchall()}

        assert "tenders" in tables

    def test_save_tenders_new(self):
        tenders = [
            dict(TENDER_DATA),
            {
                **TENDER_DATA,
                "reference": "BOAMP-2024AO002",
                "titre": "Autre mission",
            },
        ]
        count = self._db.save_tenders(tenders)
        assert count == 2

    def test_save_tenders_idempotent(self):
        """Sauvegarder les mêmes AOs deux fois → 0 doublons."""
        tenders = [dict(TENDER_DATA)]
        count1 = self._db.save_tenders(tenders)
        count2 = self._db.save_tenders(tenders)
        assert count1 == 1
        assert count2 == 0

    def test_save_tenders_partial_duplicate(self):
        """Un AO nouveau + un doublon → 1 enregistré."""
        count1 = self._db.save_tenders([dict(TENDER_DATA)])
        new = {**TENDER_DATA, "reference": "BOAMP-NEW999", "titre": "Nouveau"}
        count2 = self._db.save_tenders([dict(TENDER_DATA), new])
        assert count1 == 1
        assert count2 == 1

    def test_update_analysis(self):
        self._db.save_tenders([dict(TENDER_DATA)])
        self._db.update_analysis(TENDER_DATA["reference"], GEMINI_ANALYSIS)

        rows = self._db.get_tenders()
        assert rows[0]["decision"] == "GO"
        assert rows[0]["score"] == 82
        analyse = json.loads(rows[0]["analyse"])
        assert analyse["resume"] == GEMINI_ANALYSIS["resume"]

    def test_get_tenders_all(self):
        self._db.save_tenders([dict(TENDER_DATA)])
        rows = self._db.get_tenders()
        assert len(rows) == 1
        assert rows[0]["reference"] == TENDER_DATA["reference"]

    def test_get_tenders_filter_source(self):
        self._db.save_tenders([dict(TENDER_DATA)])
        fm = {**TENDER_DATA, "reference": "FM-001", "source": "francemarches"}
        self._db.save_tenders([fm])

        boamp_rows = self._db.get_tenders(source="boamp")
        fm_rows = self._db.get_tenders(source="francemarches")

        assert all(r["source"] == "boamp" for r in boamp_rows)
        assert all(r["source"] == "francemarches" for r in fm_rows)

    def test_get_tenders_filter_decision(self):
        self._db.save_tenders([dict(TENDER_DATA)])
        self._db.update_analysis(TENDER_DATA["reference"], GEMINI_ANALYSIS)

        nogo = {**TENDER_DATA, "reference": "BOAMP-NOGO"}
        self._db.save_tenders([nogo])
        self._db.update_analysis("BOAMP-NOGO", {**GEMINI_ANALYSIS, "decision": "NO-GO"})

        go_rows = self._db.get_tenders(decision="GO")
        nogo_rows = self._db.get_tenders(decision="NO-GO")

        assert len(go_rows) == 1
        assert len(nogo_rows) == 1

    def test_get_unanalyzed_references(self):
        self._db.save_tenders([dict(TENDER_DATA)])
        unanalyzed = self._db.get_unanalyzed_references([TENDER_DATA["reference"]])
        assert TENDER_DATA["reference"] in unanalyzed

        self._db.update_analysis(TENDER_DATA["reference"], GEMINI_ANALYSIS)
        unanalyzed_after = self._db.get_unanalyzed_references(
            [TENDER_DATA["reference"]]
        )
        assert unanalyzed_after == []

    def test_get_unanalyzed_empty_list(self):
        result = self._db.get_unanalyzed_references([])
        assert result == []

    def test_export_to_excel(self, tmp_path):
        self._db.save_tenders([dict(TENDER_DATA)])
        self._db.update_analysis(TENDER_DATA["reference"], GEMINI_ANALYSIS)

        export_path = str(tmp_path / "export.xlsx")
        self._db.export_to_excel(export_path)

        assert Path(export_path).exists()

        import openpyxl

        wb = openpyxl.load_workbook(export_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert "Titre" in headers
        assert "Décision" in headers
        assert ws.max_row >= 2

    def test_default_db_path(self):
        """Instanciation sans db_path → chemin par défaut calculé et créé (couvre lignes 18-21)."""
        from utils.tender_db import TenderDatabase

        db = TenderDatabase()  # db_path=None → branche lignes 18-21 exécutée

        assert "tenderscout" in db.db_path
        assert db.db_path.endswith("tenders.db")
        assert Path(db.db_path).exists()

    def test_export_to_excel_invalid_analyse_json(self, tmp_path):
        """Analyse avec JSON invalide dans la colonne → géré sans exception (couvre lignes 216-217)."""
        import sqlite3

        from utils.tender_db import TenderDatabase

        db_path = str(tmp_path / "invalid_json.db")
        db = TenderDatabase(db_path=db_path)
        db.save_tenders([dict(TENDER_DATA)])

        # Insérer directement un JSON invalide dans la colonne `analyse`
        with sqlite3.connect(db_path) as conn:
            conn.execute(
                "UPDATE tenders SET analyse = ? WHERE reference = ?",
                ("NOT_VALID_JSON{{{", TENDER_DATA["reference"]),
            )
            conn.commit()

        export_path = str(tmp_path / "export_invalid.xlsx")
        db.export_to_excel(export_path)  # Ne doit pas lever d'exception

        assert Path(export_path).exists()


# ============================================================
# TestTenderScoutEndpoints
# ============================================================


class TestTenderScoutEndpoints:
    """Tests pour les endpoints FastAPI TenderScout."""

    # ── Page HTML ─────────────────────────────────────────────

    def test_page_route(self):
        with patch("app.get_current_user", return_value=MOCK_USER):
            resp = client.get("/tenderscout")
        assert resp.status_code == 200
        assert "TenderScout" in resp.text

    def test_page_route_redirects_unauthenticated(self):
        with patch("app.get_current_user", return_value=None):
            resp = client.get("/tenderscout", follow_redirects=False)
        assert resp.status_code in (302, 307)

    # ── POST /api/tenderscout/scan ────────────────────────────

    def test_scan_ok(self):
        with (
            patch("app.get_current_user", return_value=MOCK_USER),
            patch("app._run_tender_scout"),
        ):
            resp = client.post(
                "/api/tenderscout/scan",
                json={"keywords": ["data"], "sources": ["boamp"]},
                headers=CSRF_HEADERS,
            )
        assert resp.status_code == 200
        assert "job_id" in resp.json()

    def test_scan_empty_keywords(self):
        with patch("app.get_current_user", return_value=MOCK_USER):
            resp = client.post(
                "/api/tenderscout/scan",
                json={"keywords": [], "sources": ["boamp"]},
                headers=CSRF_HEADERS,
            )
        assert resp.status_code == 400

    def test_scan_whitespace_keywords_only(self):
        with patch("app.get_current_user", return_value=MOCK_USER):
            resp = client.post(
                "/api/tenderscout/scan",
                json={"keywords": ["  ", ""], "sources": ["boamp"]},
                headers=CSRF_HEADERS,
            )
        assert resp.status_code == 400

    def test_scan_invalid_source(self):
        with patch("app.get_current_user", return_value=MOCK_USER):
            resp = client.post(
                "/api/tenderscout/scan",
                json={"keywords": ["test"], "sources": ["invalid_source"]},
                headers=CSRF_HEADERS,
            )
        assert resp.status_code == 400
        assert "invalides" in resp.json()["detail"]

    def test_scan_missing_sources(self):
        with patch("app.get_current_user", return_value=MOCK_USER):
            resp = client.post(
                "/api/tenderscout/scan",
                json={"keywords": ["test"]},
                headers=CSRF_HEADERS,
            )
        assert resp.status_code == 400

    def test_scan_unauthenticated(self):
        with patch("app.get_current_user", return_value=None):
            resp = client.post(
                "/api/tenderscout/scan",
                json={"keywords": ["data"], "sources": ["boamp"]},
                headers=CSRF_HEADERS,
            )
        assert resp.status_code == 401

    def test_scan_invalid_json(self):
        with patch("app.get_current_user", return_value=MOCK_USER):
            resp = client.post(
                "/api/tenderscout/scan",
                content=b"not json",
                headers={**CSRF_HEADERS, "Content-Type": "application/json"},
            )
        assert resp.status_code == 400

    # ── GET /api/tenderscout/stream/{job_id} ─────────────────

    def test_stream_not_found(self):
        resp = client.get("/api/tenderscout/stream/unknown_job_id")
        assert resp.status_code == 200
        assert "error_msg" in resp.text

    def test_stream_done(self):
        from app import jobs

        job_id = "ts_test1"
        jobs[job_id] = {
            "type": "tenderscout",
            "status": "done",
            "steps": [{"step": "done", "status": "done", "progress": 100}],
            "result": {"total": 5, "new": 3, "analyzed": 3},
            "error": None,
        }
        try:
            resp = client.get(f"/api/tenderscout/stream/{job_id}")
            assert resp.status_code == 200
            assert "result" in resp.text
        finally:
            del jobs[job_id]

    def test_stream_error(self):
        from app import jobs

        job_id = "ts_test_err"
        jobs[job_id] = {
            "type": "tenderscout",
            "status": "error",
            "steps": [],
            "result": None,
            "error": "Scraping échoué",
        }
        try:
            resp = client.get(f"/api/tenderscout/stream/{job_id}")
            assert resp.status_code == 200
            assert "error_msg" in resp.text
        finally:
            del jobs[job_id]

    # ── GET /api/tenderscout/tenders ─────────────────────────

    def test_list_tenders(self):
        with (
            patch("app.get_current_user", return_value=MOCK_USER),
            patch("app.TenderDatabase") as MockDB,
        ):
            MockDB.return_value.get_tenders.return_value = [dict(TENDER_DATA)]
            resp = client.get("/api/tenderscout/tenders")

        assert resp.status_code == 200
        data = resp.json()
        assert isinstance(data, list)
        assert len(data) == 1

    def test_list_tenders_with_filters(self):
        with (
            patch("app.get_current_user", return_value=MOCK_USER),
            patch("app.TenderDatabase") as MockDB,
        ):
            MockDB.return_value.get_tenders.return_value = []
            resp = client.get(
                "/api/tenderscout/tenders?source=boamp&decision=GO"
            )
            MockDB.return_value.get_tenders.assert_called_once_with(
                source="boamp", decision="GO"
            )

        assert resp.status_code == 200

    def test_list_tenders_unauthenticated(self):
        with patch("app.get_current_user", return_value=None):
            resp = client.get("/api/tenderscout/tenders")
        assert resp.status_code == 401

    # ── GET /api/tenderscout/export ──────────────────────────

    def test_export_excel(self, tmp_path):
        fake_xlsx = tmp_path / "fake.xlsx"
        fake_xlsx.write_bytes(b"PK")  # Fake zip header

        with (
            patch("app.get_current_user", return_value=MOCK_USER),
            patch("app.TenderDatabase") as MockDB,
            patch("app.datetime") as mock_dt,
        ):
            mock_dt.now.return_value.strftime.return_value = "20240301_120000"
            MockDB.return_value.export_to_excel.side_effect = lambda p: Path(p).write_bytes(b"PK")
            resp = client.get("/api/tenderscout/export")

        assert resp.status_code == 200

    def test_export_excel_unauthenticated(self):
        with patch("app.get_current_user", return_value=None):
            resp = client.get("/api/tenderscout/export")
        assert resp.status_code == 401

    # ── POST /api/tenderscout/notify ─────────────────────────

    def test_notify_ok(self):
        mock_service = MagicMock()
        mock_draft = {"id": "draft_abc123", "message_id": "msg_xyz"}

        with (
            patch("app.get_current_user", return_value=MOCK_USER),
            patch("app.TenderDatabase") as MockDB,
            patch("app.MeetingGmailClient") as MockGmail,
        ):
            go_tender = {**TENDER_DATA, "decision": "GO", "score": 80, "analyse": None}
            MockDB.return_value.get_tenders.return_value = [go_tender]
            MockGmail.return_value.authenticate.return_value = mock_service
            MockGmail.return_value.create_draft.return_value = mock_draft

            resp = client.post(
                "/api/tenderscout/notify",
                json={
                    "to": "user@example.com",
                    "credentials_path": "/tmp/creds.json",
                },
                headers=CSRF_HEADERS,
            )

        assert resp.status_code == 200
        assert resp.json()["id"] == "draft_abc123"

    def test_notify_no_go_tenders(self):
        """Récap avec 0 AOs GO → brouillon créé quand même."""
        mock_service = MagicMock()
        mock_draft = {"id": "draft_empty", "message_id": "msg_empty"}

        with (
            patch("app.get_current_user", return_value=MOCK_USER),
            patch("app.TenderDatabase") as MockDB,
            patch("app.MeetingGmailClient") as MockGmail,
        ):
            MockDB.return_value.get_tenders.return_value = []
            MockGmail.return_value.authenticate.return_value = mock_service
            MockGmail.return_value.create_draft.return_value = mock_draft

            resp = client.post(
                "/api/tenderscout/notify",
                json={"to": "user@example.com", "credentials_path": "/tmp/creds.json"},
                headers=CSRF_HEADERS,
            )

        assert resp.status_code == 200

    def test_notify_missing_to(self):
        with patch("app.get_current_user", return_value=MOCK_USER):
            resp = client.post(
                "/api/tenderscout/notify",
                json={"credentials_path": "/tmp/creds.json"},
                headers=CSRF_HEADERS,
            )
        assert resp.status_code == 400

    def test_notify_missing_credentials(self):
        with patch("app.get_current_user", return_value=MOCK_USER):
            resp = client.post(
                "/api/tenderscout/notify",
                json={"to": "user@example.com"},
                headers=CSRF_HEADERS,
            )
        assert resp.status_code == 400

    def test_notify_draft_creation_error(self):
        from agents.meeting_capture_agent import DraftCreationError

        with (
            patch("app.get_current_user", return_value=MOCK_USER),
            patch("app.TenderDatabase") as MockDB,
            patch("app.MeetingGmailClient") as MockGmail,
        ):
            MockDB.return_value.get_tenders.return_value = []
            MockGmail.return_value.authenticate.return_value = MagicMock()
            MockGmail.return_value.create_draft.side_effect = DraftCreationError(
                "Quota exceeded"
            )

            resp = client.post(
                "/api/tenderscout/notify",
                json={"to": "user@example.com", "credentials_path": "/tmp/creds.json"},
                headers=CSRF_HEADERS,
            )

        assert resp.status_code == 500

    def test_notify_unexpected_error(self):
        with (
            patch("app.get_current_user", return_value=MOCK_USER),
            patch("app.TenderDatabase") as MockDB,
            patch("app.MeetingGmailClient") as MockGmail,
        ):
            MockDB.return_value.get_tenders.return_value = []
            MockGmail.return_value.authenticate.side_effect = Exception("Unexpected")

            resp = client.post(
                "/api/tenderscout/notify",
                json={"to": "user@example.com", "credentials_path": "/tmp/creds.json"},
                headers=CSRF_HEADERS,
            )

        assert resp.status_code == 500

    def test_notify_unauthenticated(self):
        with patch("app.get_current_user", return_value=None):
            resp = client.post(
                "/api/tenderscout/notify",
                json={"to": "user@example.com", "credentials_path": "/tmp/creds.json"},
                headers=CSRF_HEADERS,
            )
        assert resp.status_code == 401

    def test_notify_invalid_json(self):
        with patch("app.get_current_user", return_value=MOCK_USER):
            resp = client.post(
                "/api/tenderscout/notify",
                content=b"bad json",
                headers={**CSRF_HEADERS, "Content-Type": "application/json"},
            )
        assert resp.status_code == 400
